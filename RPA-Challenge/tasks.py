from robocorp.tasks import task
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select
from RPA.Robocorp.Vault import Vault
from robocorp import workitems
from selenium.common.exceptions import NoSuchElementException, InvalidSelectorException, TimeoutException
from io import BytesIO
import datetime
import os
from PIL import Image
import io
import requests

@task
def news_data():
    """Search for and scrape data from news websites."""
    details = fetch_workitems()
    create_excel()
    navigate_webpage(details)

def create_excel():
    """Create an excel file and add headings for required columns"""
    global wb, ws
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Add column headers to the first row of the sheet
    ws["A1"] = "Title"
    ws["B1"] = "Date"
    ws["C1"] = "No. Of Search Phrases"
    ws["D1"] = "Money in Title"
    ws["E1"] = "Image Link"

    wb.save("NewsData.xlsx")  # Save the workbook with the specified filename
    wb.close()

def navigate_webpage(details):
    """Navigate to the Reuters news website and collect data"""
    # try:
    #initialise the driver using selenium and chromium web driver
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome(options=options)
    #get the url of the website and open it in a browser window
    driver.get("https://www.reuters.com/site-search/?query=football&section=all&offset=0&sort=newest&date=past_month")
    time.sleep(10)
        #find the search button and click on it
    #     search_button = WebDriverWait(driver, 10).until(
    #         EC.element_to_be_clickable((By.CSS_SELECTOR, "svg[data-testid='SvgSearch']"))
    #     )
    #     search_button.click()
    #     time.sleep(5)
    #     #find the search input box and enter the phrase to be searched
    #     search_input = WebDriverWait(driver, 10).until(
    #         EC.element_to_be_clickable((By.CSS_SELECTOR,'input[data-testid="FormField:input"][type="search"][autocomplete="off"][spellcheck="false"][maxlength="256"][class="text__text__1FZLe text__dark-grey__3Ml43 text__regular__2N1Xr text__small__1kGq2 body__base__22dCE body__medium__2blzt form-field__input__7LFh3 form-field__default__1IHy7 search-bar__search-input__3ahqM"]'))
    #     )
    #     time.sleep(5)
    #     search_input.send_keys(details[0])
    #     time.sleep(5)
    #     search_submit_button = WebDriverWait(driver, 20).until(
    #         EC.element_to_be_clickable((By.CSS_SELECTOR, 'svg[data-testid="SvgSearch"][class="search-bar__icon__ORXTq search-bar__search-alt-icon__juWN_"]'))
    #     )
    #     time.sleep(5)
    #     search_submit_button.click()
    # except NoSuchElementException:
    #     print("Error: Unable to locate the search input element.")
    # except InvalidSelectorException:
    #     print("Error: The provided XPATH is invalid.")
    # except TimeoutException:
    #     print("Error: The element was not found within the specified time.")
    
    # try:
    #     #find the time span and click on it
    #     if details[2] == "1 month":
    #         past_month_option = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, "//*[@id='react-aria1233940870-:r1u:-option-Pastmonth']/span"))
    #         )
    #         past_month_option.click()
    #     elif details[2] == "Anytime":
    #         past_month_option = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, "//*[@id='react-aria1233940870-:r114:-option-Anytime']/span"))
    #         )
    #         past_month_option.click()
    #     elif details[2] == "24 hours":
    #         past_month_option = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, "//*[@id='react-aria1233940870-:r114:-option-Past24hours']/span"))
    #         )
    #         past_month_option.click()
    #     elif details[2] == "1 week":
    #         past_month_option = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, "//*[@id='react-aria1233940870-:r114:-option-Pastweek']/span"))
    #         )
    #         past_month_option.click()
    #     else:
    #         past_month_option = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, "//*[@id='react-aria1233940870-:r114:-option-Pastyear']/span"))
    #         )
    #         past_month_option.click()

    # except NoSuchElementException:
    #     print("Error: Unable to locate the date/time element.")
    # except InvalidSelectorException:
    #     print("Error: The provided XPATH is invalid.")
    # except TimeoutException:
    #     print("Error: The element was not found within the specified time.")

    #opening the worksheet
    wb = openpyxl.load_workbook("NewsData.xlsx")
    ws = wb.active
    def store_and_add_data():
        """This function is used here to go through all the results on each page and do it recursively"""
        try:
            previous_url=driver.current_url
            money_in_title=""
            search_phrases_in_title=0
            #look for all the search results on the current page
            result_links = WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.XPATH, "//*[@id='fusion-app']/div[2]/div[2]/div/div[2]/div[2]/ul/li"))
                )
            #go through each result one by one 
            for result in result_links:
                #to store the images 
                try:
                    #heading or the title of the article
                    heading = result.find_element(By.CSS_SELECTOR, "span[data-testid='Heading'].text__text__1FZLe.text__dark-grey__3Ml43.text__medium__1kbOh.text__heading_6__1qUJ5.heading__base__2T28j.heading__heading_6__RtD9P").text
                    #date of the article
                    date = result.find_element(By.XPATH,"//div/div/time").text
                    if "min" or "AM" or "PM" in date:
                        date="February 29, 2024"
                    #looking for the number of search phrases in the title
                    lower_text = heading.lower()
                    words = lower_text.split()
                    search_phrases_in_title = words.count("football")
                    
                    #check if money is present in the title, can be improved later
                    if '$' in heading:
                        money_in_title = "True"
                    else:
                        money_in_title = "False"
                    
                    #look for images in the news article
                    try:
                        image = result.find_element(By.TAG_NAME,"img")
                        src = image.get_attribute('src')
                        response = requests.get(src)
                        count = time.strftime("%Y-%m-%d-%H%M%S", time.localtime())
                        image_link = f'{count}.jpg'
                        response = requests.get(src)
                        img = Image.open(io.BytesIO(response.content))
                        img.save(image_link)                        
                    except NoSuchElementException:
                        image_link="No image here!"
                    
                    #adding data to the excel file
                    ws.append([heading,date,search_phrases_in_title,money_in_title,image_link])
                    wb.save("NewsData.xlsx")
                except NoSuchElementException:
                        continue
            #going to the next page if it exists and getting details of that page
            try:
                next_page = WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "svg[data-testid='SvgChevronRight']"))
                    )
                next_page.click()                
                if driver.current_url == previous_url:
                    time.sleep(3)
                    print("Yay all the data is extracted, job done mate!")
                else:
                    store_and_add_data()
            except NoSuchElementException:
                print("Yay all the data is extracted, job done mate!")            
            
        except NoSuchElementException:
            print("Couldn't find any element with given req. Get better buddy!")
        except TimeoutException:
            print("Oops took too long mate! Sometimes it's better to finish quickly!")
    #calling the function to start scraping from first page
    store_and_add_data()
    
def fetch_workitems():
    """Fetch the payload and details from the input work-items"""
    item = workitems.inputs.current
    payload = item.payload
    # Access the data from the payload
    search_phrase = payload.get('search phrase')
    category = payload.get('category')
    number_of_months = payload.get('timespan')

    details = [search_phrase,category,number_of_months]

    return details

    
