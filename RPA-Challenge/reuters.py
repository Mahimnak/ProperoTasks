import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from io import BytesIO
import datetime
import os
from PIL import Image
import io
import requests
from config import *
from logger import setup_logger

class Reuters:
    def __init__(self, work_item) -> None:
        """
        Initialize the AljazeeraNews class.
        """
        self.search_term = work_item["search phrase"]
        self.category = work_item["category"]
        self.number_of_months = work_item["timespan"]
        log_file_path = os.path.join(LOGS_DIR, LOG_FILE)
        self.logger = setup_logger(log_file_path)    

    def create_excel(self)-> None:
        """Create an excel file and add headings for required columns"""
        global wb, ws
        # Create a workbook and select the active worksheet
        self.logger.info("Creating Excel file...")
        wb = Workbook()
        ws = wb.active
        
        # Add column headers to the first row of the sheet
        ws["A1"] = "Title"
        ws["B1"] = "Date"
        ws["C1"] = "No. Of Search Phrases"
        ws["D1"] = "Money in Title"
        ws["E1"] = "Image Link"

        wb.save("output\excel\NewsData.xlsx")  # Save the workbook with the specified filename
        wb.close()

    def navigate_webpage(self)-> None:
        """Navigate to the Reuters news website and collect data"""
        self.logger.info("Navigating to Reuters news website...")
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        driver = webdriver.Chrome(options=options)
        #get the url of the website and open it in a browser window
        driver.get("https://www.reuters.com/site-search/?query=football&section=all&offset=0&sort=newest&date=past_month")
        time.sleep(10)    
        wb = openpyxl.load_workbook("NewsData.xlsx")
        ws = wb.active
        def store_and_add_data():
            """This function is used here to go through all the results on each page and do it recursively"""
            try:
                previous_url=driver.current_url
                money_in_title=""
                search_phrases_in_title=0
                #look for all the search results on the current page
                result_links = WebDriverWait(driver, 20).until(
                        EC.presence_of_all_elements_located((By.XPATH, "//*[@id='fusion-app']/div[2]/div[2]/div/div[2]/div[2]/ul/li"))
                    )
                #go through each result one by one 
                for result in result_links:
                    #to store the images 
                    try:
                        #heading or the title of the article
                        heading = result.find_element(By.CSS_SELECTOR, "span[data-testid='Heading'].text__text__1FZLe.text__dark-grey__3Ml43.text__medium__1kbOh.text__heading_6__1qUJ5.heading__base__2T28j.heading__heading_6__RtD9P").text
                        #date of the article
                        date = result.find_element(By.XPATH,"//div/div/time").text
                        if "min" or "AM" or "PM" in date:
                            date="February 29, 2024"
                        #looking for the number of search phrases in the title
                        lower_text = heading.lower()
                        words = lower_text.split()
                        search_phrases_in_title = words.count("football")
                        
                        #check if money is present in the title, can be improved later
                        if '$' in heading:
                            money_in_title = "True"
                        else:
                            money_in_title = "False"
                        
                        #look for images in the news article
                        try:
                            image = result.find_element(By.TAG_NAME,"img")
                            src = image.get_attribute('src')
                            response = requests.get(src)
                            count = time.strftime("%Y-%m-%d-%H%M%S", time.localtime())
                            image_link = f'{count}.jpg'
                            response = requests.get(src)
                            img = Image.open(io.BytesIO(response.content))
                            img.save(f"output\images\{image_link}") 

                        except NoSuchElementException:
                            self.logger.error("Exception occurred while finding an image tag")
                            image_link="No image here!"
                        
                        #adding data to the excel file
                        ws.append([heading,date,search_phrases_in_title,money_in_title,image_link])
                        wb.save("NewsData.xlsx")
                    except NoSuchElementException:
                            self.logger.error("Exception occurred while finding a row")
                            continue
                #going to the next page if it exists and getting details of that page
                try:
                    next_page = WebDriverWait(driver, 20).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "svg[data-testid='SvgChevronRight']"))
                        )
                    next_page.click()                
                    if driver.current_url == previous_url:
                        time.sleep(3)
                        
                    else:
                        store_and_add_data()
                except NoSuchElementException:
                    self.logger.info("Well fortunately all the data was extracted so no more data to be extracted!")
                                
            except NoSuchElementException:
                self.logger.error("Couldn't find any element with given req. Get better buddy!")
            except TimeoutException:
                self.logger.error("Couldn't find any element with given req. Get better buddy!")
        
        #calling the function to start scraping from first page
        store_and_add_data()