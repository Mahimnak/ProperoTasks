import unicodedata
import re
import fitz
import requests
import tempfile
import os
import openpyxl
import pandas as pd

class pdf_data:
    def __init__(self,pdf_path) -> None:
        self.pdf_path = pdf_path

    def remove_non_ascii(self, text)->str:
        """Clean the text extracted from the table and remove any unwanted characters."""
        cleaned_text = re.sub(r'[^\x00-\x7F]', '', text)
        cleaned_text = re.sub(r'[\u263a-\U0001f645]', '', cleaned_text)
        cleaned_text = cleaned_text.replace('/', '')
        return cleaned_text.strip()

    # def download_pdf(self, pdf_url, save_path):
    #     response = requests.get(pdf_url)
    #     if response.status_code == 200:
    #         with open(save_path, 'wb') as f:
    #             f.write(response.content)
    #     else:
    #         raise Exception("Failed to download PDF from the provided URL")


    def extract_tables_from_pdf(self)->list:
        """Extract the table from the pdf one row at a time and return the data."""
        tables_data = []
        with fitz.open(self.pdf_path) as pdf:
            page = pdf[0]
            tables = page.find_tables()

            for table in tables:
                # Extract table data
                table_data = table.extract()
                cleaned_table_data = [[self.remove_non_ascii(cell) if cell else None for cell in row] for row in table_data]
                # Convert table data to DataFrame
                df = pd.DataFrame(cleaned_table_data[1:], columns=cleaned_table_data[0])
                # # Append DataFrame to list
                tables_data.append(df)
        return tables_data

    def add_link_excel(self, link)->None:
        """After getting the link to the technical specification pdf add it to the excel file."""
        workbook = openpyxl.load_workbook("tables_from_pdf.xlsx")
        worksheet = workbook.get_sheet_by_name("bid_details")

        max_row = worksheet.max_row+1
        worksheet.cell(row =  max_row, column = 1).value = "Technical specification"
        worksheet.cell(row =  max_row, column = 2).value = link
        workbook.save("tables_from_pdf.xlsx")
        workbook.close()

        
    def save_tables_to_excel(self, tables, worksheet_name)->None:
        """Add the table rows to the excel file."""
        workbook = openpyxl.load_workbook("tables_from_pdf.xlsx")
        worksheet = workbook.get_sheet_by_name(worksheet_name)
        for idx, table_df in enumerate(tables):
            for r_idx, (_, row) in enumerate(table_df.iterrows(), start=1):
                for c_idx, value in enumerate(row, start=1):
                    value = str(value)
                    # Replace any characters that cannot be used in Excel worksheets
                    value = ''.join(char if unicodedata.category(char)[0] != 'C' else '' for char in value)
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
        workbook.save("tables_from_pdf.xlsx")

    def extract_hyperlinks_from_pdf(self)->str:
        """Get the link to the technical specification pdf and return it as a string."""
        hyperlinks = []
        with fitz.open(self.pdf_path) as pdf:
            page = pdf[2]
            links = page.get_links()
            for link in links:
                hyperlink = {
                    "page_number": 3,
                    "url": link.get('uri', ''),
                    "rect": link.get('rect', [])
                }
                hyperlinks.append(hyperlink)
        tech_specs = hyperlinks[1]["url"]
        return tech_specs



  
    
def main():
    pdf_path = "https://mkp.gem.gov.in/uploaded_documents/51/16/877/OrderItem/BoqDocument/2024/1/20/gem_nit-final_2024-01-20-11-42-11_976251cc06053ae9af6f3014a0a36c40.pdf"
    pdf = pdf_data("GeM-Bidding-5927594.pdf")
    pdf.text_data_test()
    # table = pdf.extract_tables_from_pdf()
    # pdf.save_tables_to_excel(table, "technical_specification")

    # pdf_filename = os.path.basename(pdf_path)
    # save_path = os.path.join(os.path.dirname(__file__), pdf_filename)  # Save in the same directory as the program
    # pdf.download_pdf(pdf_path,save_path)

if __name__ == "__main__":
    main()




# def text_data_test(self):
    #     pdf = fitz.open(self.pdf_path)
    #     # Iterate over each page
    #     for page in pdf:
    #         # Get the text of the page
    #         page_text = page.get_text()
    #         # Split the text into lines
    #         lines = page_text.split("\n")
    #         # Iterate over each line
    #         for line in lines:
    #             # Do something with the line
    #             print(line)