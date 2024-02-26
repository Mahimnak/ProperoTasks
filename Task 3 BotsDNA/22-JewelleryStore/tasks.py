from robocorp.tasks import task
from robocorp import browser
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

@task
def jewellery_store():
    
    engagement()
    Men_Jewelry()
    Women_Jewelry()
    
def Men_Jewelry():
    workbook = openpyxl.load_workbook("CurrentMonthNewJewelry.xlsx")
    sheet = workbook.get_sheet_by_name("Men_Jewelry")
   
    metal_type = []
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Bracelet":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver = webdriver.Chrome()
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Earring":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Necklace":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()  
    driver.quit()

def Women_Jewelry():
    workbook = openpyxl.load_workbook("CurrentMonthNewJewelry.xlsx")
    sheet = workbook.get_sheet_by_name("Women_Jewelry")
   
    metal_type = []
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Bracelet":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver = webdriver.Chrome()
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Earring":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Necklace":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()  
    driver.quit()
     
def engagement():
    workbook = openpyxl.load_workbook("CurrentMonthNewJewelry.xlsx")
    sheet = workbook.get_sheet_by_name("Engagement")
    metal_type = []
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Ring":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver = webdriver.Chrome()
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Bracelet":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Earring":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()

    metal_type=[]
    metal_temp=""
    for row_num, cell in enumerate(sheet['A'], start=1):
            if cell.value == "Necklace":
                row_value = row_num
                metal_temp = sheet.cell(row = row_value, column=2).value
                if metal_temp not in metal_type:
                     metal_type.append(metal_temp)
    driver.get("https://botsdna.com/jewelry/")

    driver.find_element(By.XPATH,"//*[@id='ddlCategory']/option[2]").click()
    driver.find_element(By.XPATH,"//*[@id='ddlJewelry']/option[2]").click()
    for i in range(0,len(metal_type)):
         td = driver.find_element(By.XPATH,f"//td[contains(., {metal_type[i]})]")
         td.find_element(By.TAG_NAME,"input").click()
    driver.find_element(By.XPATH,"//*[@id='courts']/tbody/tr[4]/td/input").click()  
    driver.quit()