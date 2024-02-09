from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from robocorp import vault
from RPA.Excel.Application import Application
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException


@task
def locator():
    """Locating the company and country information on a website and adding it to an excel file."""
    wb = Workbook()
    ws = wb.active
    driver = webdriver.Chrome()
    new_elements=[]
    url="https://botsdna.com/locator/"
    driver.get(url)
    wait = WebDriverWait(driver, 10)
    # elements = driver.find_elements(By.XPATH, "//*[contains(@id, 'tbl')]/tbody/tr/td[1]")
    elements = driver.find_elements(By.XPATH, "//*[contains(@id, 'tbl')]/tbody/tr")
    abc = driver.find_elements(By.XPATH, "//*[contains(@id, 'tbl')]/tbody/tr[1]/th")
    b=[]
    for a in abc:
       if a.text!="":
            b.append(a.text)
    for i in range(1,len(b)):
        sheet = wb.create_sheet(b[i])
        sheet.title = b[i]
        sheet["A1"] = "CustomerName"
        sheet["B1"] = "Number of Locations"

    wb.save("output.xlsx")
    for i in range(0,len(elements)):
        if elements[i].text!="":
            new_elements.append(elements[i])
    
    for i in range(0,len(new_elements)):
        if i==0:
            continue

        inner_elements =  new_elements[i].find_elements(By.TAG_NAME,"td")
        for j in range(1,len(inner_elements)):
            
            if inner_elements[j].text!=0:
                sheet = wb[b[j]]
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=inner_elements[0].text)
                sheet.cell(row=row, column=2, value=inner_elements[j].text)
    
    wb.save("output.xlsx")



    