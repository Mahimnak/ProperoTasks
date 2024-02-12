from robocorp.tasks import task
from robocorp import browser
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import io
import zipfile
import time
from selenium.webdriver.support.relative_locator import locate_with

@task
def active_loans():
    """Get details of customer and store it in the input.xlsx file"""
    driver = webdriver.Chrome()
    url="https://botsdna.com/ActiveLoans/"
    driver.get(url)
    acc = ""

    for i in range(2,79):
        acc = get_accountno(i)
        last_digits_acc = acc[-4:]
        a_tags = driver.find_elements(By.XPATH,f"//a[contains(., '{last_digits_acc}')]")
        if len(a_tags)>0:
            a_tags[0].click()
            time.sleep(1)
            open_text_add_data(acc,i)
            row_element = driver.find_element(By.XPATH,f"//tr[.//td/a[contains(., '{last_digits_acc}')]]")
            pan = row_element.find_element(By.XPATH,"./td[3]")
            status = row_element.find_element(By.XPATH,"./td[1]")
            
            workbook = openpyxl.load_workbook('input.xlsx')
            sheet = workbook.active
            
            sheet.cell(row=i, column=7, value = pan.text)
            sheet.cell(row=i, column=8, value = status.text)

            workbook.save('input.xlsx')
            workbook.close()

def get_accountno(a):
    """Get the account number from the input.xlsx file"""
    workbook = openpyxl.load_workbook('input.xlsx')
    sheet = workbook.active
    # Get the value in the first row of the first column
    first_cell_value = sheet.cell(row=a, column=1).value
    workbook.close()
    return  first_cell_value

def open_text_add_data(account_number, a):
    """Open the downloaded text file and copy-paste the data in the input.xlsx file"""
    with zipfile.ZipFile(f"C:/Users/Mahimna/Downloads/{account_number}.zip", 'r') as zip_file:
    # Extract the text file from the zip file
        with zip_file.open(f"{account_number}.txt") as text_file:
            # Read the contents of the text file
            contents = text_file.read().decode('utf-8')
    content_list = contents.split('\n')
    content_list = [x.strip() for x in contents.split('\n') if x]
    #['Account Number: 978503257325', '', 'Bank: SAPTHAGIRI GRAMIN BANK', '', 'Branch: RAICHUR', '', 'Loan Taken On: 30-NOV-2007', '', 'Amount: 3400000', '', 'EMI(month): 23631']

    bank = content_list[2][len('Bank:'):]
    branch = content_list[4][len('Branch:'):]
    loan = content_list[6][len('Loan Taken On:'):]
    amount = content_list[8][len('Amount:'):]
    emi = content_list[10][len('EMI(month):'):]
    
    workbook = openpyxl.load_workbook('input.xlsx')
    sheet = workbook.active

    sheet.cell(row=a, column=2, value = bank)
    sheet.cell(row=a, column=3, value = branch)
    sheet.cell(row=a, column=4, value = loan)
    sheet.cell(row=a, column=5, value = amount)
    sheet.cell(row=a, column=6, value = emi)

    workbook.save('input.xlsx')
    workbook.close()

