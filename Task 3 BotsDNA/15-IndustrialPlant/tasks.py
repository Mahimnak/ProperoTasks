from robocorp.tasks import task
from robocorp import browser
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select

@task
def industrial_plant():
    """Copy data and paste it in the Master.xlsx"""
    driver = webdriver.Chrome()
    url="https://botsdna.com/IndustrialPlants/"
    driver.get(url)
    for i in range(1,4):
        equipment = driver.find_element(By.XPATH,f"//*[@id='ddlEquipment']/option[{i}]")        
        equipment.click()
        for j in range(1,4):
            plant = driver.find_element(By.XPATH,f"//*[@id='ddlPlant']/option[{j}]")
            plant.click()
            if i==1:
                rows = driver.find_elements(By.XPATH,"//*[@id='Plant1_Belts']/tbody/tr")
                for k in range(0,len(rows)):
                    if k==0:
                        pass
                    else:
                        data = rows[k].find_elements(By.TAG_NAME,"td")
                        workbook = openpyxl.load_workbook('Master.xlsx')
                        sheet = workbook.active
                        max_row = sheet.max_row+1
                        sheet.cell(row=max_row,column=1,value = data[0].text)
                        sheet.cell(row=max_row,column=5,value = data[1].text)
                        sheet.cell(row=max_row,column=4,value = data[2].text)
                        workbook.save('Master.xlsx')
                        workbook.close
            elif i==2:
                rows = driver.find_elements(By.XPATH,"//*[@id='Plant1_Tanks']/tbody/tr")
                for k in range(0,len(rows)):
                    if k==0:
                        pass
                    else:
                        data = rows[k].find_elements(By.TAG_NAME,"td")
                        workbook = openpyxl.load_workbook('Master.xlsx')
                        sheet = workbook.active
                        max_row = sheet.max_row+1
                        sheet.cell(row=max_row,column=1,value = data[0].text)
                        sheet.cell(row=max_row,column=2,value = data[1].text)
                        sheet.cell(row=max_row,column=3,value = data[2].text)
                        workbook.save('Master.xlsx')
                        workbook.close
            else:
                rows = driver.find_elements(By.XPATH,"//*[@id='Plant1_Ovens']/tbody/tr")
                for k in range(0,len(rows)):
                    if k==0:
                        pass
                    else:
                        data = rows[k].find_elements(By.TAG_NAME,"td")
                        workbook = openpyxl.load_workbook('Master.xlsx')
                        sheet = workbook.active
                        max_row = sheet.max_row+1
                        sheet.cell(row=max_row,column=1,value = data[0].text)
                        sheet.cell(row=max_row,column=6,value = data[1].text)
                        workbook.save('Master.xlsx')
                        workbook.close

            

