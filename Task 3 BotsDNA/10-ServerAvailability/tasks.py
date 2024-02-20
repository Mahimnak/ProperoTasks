from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
import openpyxl
from openpyxl import Workbook
import os
import sys
import shutil
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import zipfile
from selenium.webdriver.support.select import Select

@task
def ServerAvailability():
    """Check whether a server is available or not"""
    url="https://botsdna.com/ServerAvailability/"
    driver = webdriver.Chrome()
    for i in range(2,37):
        driver.get(url)
        workbook = openpyxl.load_workbook('input.xlsx')
        sheet = workbook.active 
        user_id = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        user_input = driver.find_element(By.ID,"username")
        user_input.send_keys(user_id)
        pass_input = driver.find_element(By.ID, "password")
        pass_input.send_keys(password)
        server = driver.find_element(By.XPATH,f"//*[@id='name']/option[{i}]")
        server.click()
        submit = driver.find_element(By.XPATH,"//table/tbody/tr[4]/td[2]/input")
        submit.click()
        status = driver.find_element(By.XPATH,"//*[@id='status']").text
        status = status.splitlines()[1]
        sheet.cell(row=i, column=5, value = status)
        workbook.save("input.xlsx")
