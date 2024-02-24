from robocorp.tasks import task
from robocorp import browser
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select
import os
import io

@task
def book_shop():
    """Create excel sheet and update book details and monthly sales"""
    wb = openpyxl.load_workbook("Monthly_report.xlsx")
    
    folder_path = 'O:\\RPA\\Task 3 BotsDNA\\17-BookShop\\Monthly Data'
    for filename in os.listdir(folder_path):
        if filename.endswith('.txt'):
            file_path = os.path.join(folder_path, filename)
            with io.open(file_path, 'r', encoding='utf-8') as file:
                file_name = filename
                for i, line in enumerate(file, start=1):
                    if "|" not in line or line.count("|") < 1:  # Skip lines without enough "|" characters
                        continue
                    book_details = line.strip().split("|")

                    if i < 5:
                        continue

                    if book_details[0] not in wb.sheetnames:
                        new_sheet = wb.create_sheet(book_details[0])
                        new_sheet.cell(row=1, column=1, value="Shop Name")
                        new_sheet.cell(row=1, column=2, value="No of Books Sold")
                        new_sheet.cell(row=1, column=3, value="Month")
                        new_sheet.cell(row=2, column=1, value=file_name)
                        new_sheet.cell(row=2, column=2, value=book_details[1])
                        new_sheet.cell(row=2, column=3, value="February, 2024")
                        wb.save("Monthly_report.xlsx")
                    else:
                        sheet = wb.get_sheet_by_name(book_details[0])
                        row_value = sheet.max_row+1
                        sheet.cell(row=row_value,column=1, value= file_name)
                        sheet.cell(row=row_value,column=2, value= book_details[1])
                        sheet.cell(row=row_value, column=3, value="February, 2024")
                        wb.save("Monthly_report.xlsx")
    
    driver = webdriver.Chrome()
    driver.get("https://botsdna.com/BookShop/")

    time.sleep(10)
    
    driver.quit()
