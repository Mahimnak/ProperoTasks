from robocorp.tasks import task
from robocorp import browser
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select

@task
def courts():
    """assign courts to the teams"""
    driver = webdriver.Chrome()
    driver.get("https://botsdna.com/FootballCourtBooking/Court%20Request.html")
    
    team_rows = driver.find_elements(By.XPATH,"//*[@id='courts']/tbody/tr")
    for i in range(1,2):
        team_details = team_rows[i].find_elements(By.TAG_NAME,"td")
        req_id  = team_details[1].text
        hours  = team_details[2].text
        team_code = team_details[3].text
        driver.get("https://botsdna.com/FootballCourtBooking/Assign%20Courts.html")
        time.sleep(0.5)
        driver.find_element(By.ID,"TeamCode").send_keys(team_code)
        if i<24:
            driver.find_element(By.XPATH,f"//*[@id='CourtName']/option[{i}]").click()
        else:
            driver.find_element(By.XPATH,f"//*[@id='CourtName']/option[{i-24}]").click()
        time.sleep(0.5)
        if int(hours)==1:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[1]").click()
        elif int(hours)==2:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[2]").click()
        elif int(hours)==3:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[3]").click()
        elif int(hours)==4:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[4]").click()
        elif int(hours)==5:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[5]").click()
        elif int(hours)==6:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[6]").click()
        elif int(hours)==7:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[7]").click()
        else:
            driver.find_element(By.XPATH,f"//*[@id='NoOfHours']/option[8]").click()
        time.sleep(2)
        driver.find_element(By.XPATH,"//table/tbody/tr[4]/td/input").click()
        booking_num = driver.find_element(By.ID,"book").text
        check_in = driver.find_element(By.ID,"demo").text
        check_out = driver.find_element(By.ID,"demo1").text
        status = driver.find_element(By.XPATH,"/html/body/center[2]/table/tbody/tr[4]/td[2]/b").text
        
        wb=openpyxl.load_workbook("input.xlsx")
        sheet = wb.active

        max_row =  sheet.max_row+1
        sheet.cell(row=max_row, column=1, value=req_id)
        sheet.cell(row=max_row, column=2, value=booking_num)
        sheet.cell(row=max_row, column=3, value=check_in)        
        sheet.cell(row=max_row, column=4, value=check_out)        
        sheet.cell(row=max_row, column=5, value=status)        
        
        wb.save("input.xlsx")




