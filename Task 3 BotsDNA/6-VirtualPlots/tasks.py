from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select
from RPA.Email.ImapSmtp import ImapSmtp

@task
def virtual_plots():
    """Enter plot details, recieve booking number and all the details and send emails to respective parties"""
    for i in range(4,20):
        details = get_details(i)
        final_details = input_details(details,i)
        send_email(final_details)

def get_details(row_value):
    """Get the details of seller and buyer from the excel file"""
    workbook = openpyxl.load_workbook('input.xlsx')
    sheet = workbook.active

    seller_number = sheet.cell(row=row_value, column=1).value
    buyer_number = sheet.cell(row=row_value, column=3).value
    seller_email = sheet.cell(row=row_value, column=2).value
    buyer_email = sheet.cell(row=row_value, column=4).value
    plot_no = sheet.cell(row = row_value, column=5).value
    sqft = sheet.cell(row = row_value, column=6).value
    
    details = [seller_number,buyer_number,seller_email,buyer_email,plot_no,sqft]

    return details
    
def input_details(details,row_value):
    """Input the required details in the website"""
    url = "https://botsdna.com/vitrualplots/"
    driver = webdriver.Chrome()
    driver.get(url)
    
    row = driver.find_element(By.XPATH, f"//tr[.//td[4][contains(text(), '{details[0]}')]]")
    seller_name = row.find_element(By.XPATH, f".//td[3]").text
    radio_seller  = row.find_element(By.XPATH, ".//td/input[@name='seller']")
    radio_seller.click()

    row = driver.find_element(By.XPATH, f"//tr[.//td[4][contains(text(), '{details[1]}')]]")
    buyer_name = row.find_element(By.XPATH, f".//td[3]").text
    radio_buyer  = row.find_element(By.XPATH, ".//td/input[@name='Buyer']")
    radio_buyer.click()
    
    plot_input = driver.find_element(By.XPATH, "//table[2]/tbody/tr[1]/td[2]/input")
    plot_input.send_keys(details[4])

    sqft_input = driver.find_element(By.XPATH, "//table[2]/tbody/tr[2]/td[2]/input")
    sqft_input.send_keys(details[5])

    submit = driver.find_element(By.XPATH, "//table[2]/tbody/tr[3]/td[2]/input")
    submit.click()

    #add a wait time to load the new url
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))
    transaction_number = driver.find_element(By.XPATH,"//*[@id='TransNo']").text
    
    details.append(seller_name)
    details.append(buyer_name)
    details.append(transaction_number)

    workbook = openpyxl.load_workbook('input.xlsx')
    sheet = workbook.active

    sheet.cell(row = row_value, column = 7, value = transaction_number)
    workbook.save("input.xlsx")

    return details

def send_email(details):
    """Send emails to the respective party using the details extracted above"""
    gmail_account = details[2]
    gmail_password = "APP_PASSWORD"

    mail = ImapSmtp(smtp_server="smtp.gmail.com", smtp_port=587)
    mail.authorize(account=gmail_account, password=gmail_password)
    mail.send_message(
        sender=gmail_account,
        recipients=details[3],
        subject="Regarding plot purchase",
        body=f"seller name:{details[6]}   buyer name:{details[7]}    booking number:{details[8]}   seller number:{details[0]}   buyer number:{details[1]}",
    )