from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select

@task
def school_database():
    """Enter school codes and collect school data"""
    
    for i in range(2,32):        
        school_code = get_code(i)
        enter_code_get_details(school_code,i)

def get_code(row):
    """Get school code from Master Template.xlsx"""
    workbook = openpyxl.load_workbook('Master Template.xlsx')
    sheet = workbook.active
    
    school_code = sheet.cell(row=row, column=1).value
    return school_code

def enter_code_get_details(school_code,row_value):
    url=f"https://botsdna.com/school/{school_code}.html"
    driver = webdriver.Chrome()
    driver.get(url)
    data=[]
    # input_field = driver.find_element(By.ID,"SchoolCode")
    # input_field.send_keys(school_code)
    # submit_button = driver.find_element(By.XPATH,"//*[@id='SearchSchool']")
    # submit_button.click() 
    #add wait time to allow the page to load
    # time.sleep(5)
    school_name = driver.find_element(By.TAG_NAME,"h1").text
    data.append(school_name)
    rows = driver.find_elements(By.XPATH,"//table/tbody/tr")
    for row in rows:
        row_data = row.find_elements(By.TAG_NAME, "td")
        data.append(row_data[1].text)
    workbook = openpyxl.load_workbook('Master Template.xlsx')
    sheet = workbook.active

    for i in range(2,22):
        sheet.cell(row=row_value, column=i, value = data[i-2])
        workbook.save('Master Template.xlsx')