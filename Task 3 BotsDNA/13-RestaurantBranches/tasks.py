from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select

@task
def restaurant_branches():
    """Add restaurant details in an excel sheet and delete any duplicates"""
    driver = webdriver.Chrome()
    url = "https://botsdna.com/RestaurantBranches/3.html"
    driver.get(url)
    rows = driver.find_elements(By.XPATH,"//table/tbody/tr")
    duplicate_checker = []
    data = []
    for i in range(2,len(rows)):
        row_data = rows[i].find_elements(By.TAG_NAME,"td")
        rest_name = row_data[0].text
        gen_man = row_data[1].text
        phone  = row_data[2].text
        branch_name  = row_data[3].text
        data = [rest_name,gen_man,phone]
        if data in duplicate_checker:
            workbook = openpyxl.load_workbook("restaurant_details.xlsx")
            worksheet = workbook.active
            for row in range(1, worksheet.max_row + 1):  
                cell_value2 = worksheet.cell(row=row, column=2).value
                cell_value1 = worksheet.cell(row=row, column=1).value
                if cell_value1 ==  rest_name and cell_value2 == gen_man:
                    branch = worksheet.cell(row=last_row, column=4).value
                    branch = branch+","+branch_name
                    worksheet.cell(row=last_row, column=4, value=branch)
                else:
                    continue
            workbook.save("restaurant_details.xlsx")
            workbook.close()
        else:
            duplicate_checker.append(data)
            workbook = openpyxl.load_workbook("restaurant_details.xlsx")
            worksheet = workbook.active
            last_row = worksheet.max_row + 1
            # append data to the last empty row
            worksheet.cell(row=last_row, column=1, value=rest_name)
            worksheet.cell(row=last_row, column=2, value=gen_man)
            worksheet.cell(row=last_row, column=3, value=phone)
            worksheet.cell(row=last_row, column=4, value=branch_name)
            # save the workbook
            workbook.save("restaurant_details.xlsx")
            workbook.close()
    
    