from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
from RPA.Email.ImapSmtp import ImapSmtp
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select

@task
def server_setup():
    """Entering specification details and setting up a server"""
    for i in range(2,5):
        options = get_data(i)
        fill_data(options)

def get_data(row_value):
    """Get the data from the excel file"""
    wb = openpyxl.load_workbook("input.xlsx")
    sheet = wb.active
    os = sheet.cell(row=row_value, column=2).value
    ram = sheet.cell(row=row_value, column=3).value
    hdd = sheet.cell(row=row_value, column=4).value
    applications = sheet.cell(row=row_value, column=5).value
    values = [os,ram,hdd]
    values.append(applications.split(','))
    wb.close()
    return  values
    
def fill_data(data):
        url = "https://botsdna.com/server/"
        driver = webdriver.Chrome()
        driver.get(url)
        ram_option_value = 0
        if data[1]=="8 GB":
             ram_option_value = 1
        elif data[1]=="16 GB":
             ram_option_value = 2
        elif data[1]=="32 GB":
             ram_option_value = 3
        elif data[1]=="64 GB":
             ram_option_value = 4 
        print(ram_option_value)
        time.sleep(1)          
        select_field = driver.find_element(By.ID, "os")
        select = Select(select_field)
        select.select_by_visible_text(data[0])
        select_field = driver.find_element(By.ID, "Ram")
        select = Select(select_field)
        select.select_by_index(ram_option_value)
        time.sleep(1)
        # hdd_size  = driver.find_element(locate_with(By.TAG_NAME, "input").to_left_of({By.XPATH: f"//label[contains(., '500 GB')]"}))
        # print(hdd_size)
        # for i in range(0,len(options[3])):
        #     application = (locate_with(By.TAG_NAME, "input").to_left_of({By.XPATH, f"//label[text()={options[3][i]}]"}))
        #     applications = driver.find_element(list(application))
        #     applications.click()
        # submit_button = driver.find_element(By.XPATH,"//*[@id='CreateServer']")
        # submit_button.click()
