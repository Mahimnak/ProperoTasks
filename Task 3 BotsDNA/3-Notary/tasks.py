from robocorp.tasks import task
from robocorp import browser
from RPA.HTTP import HTTP
from RPA.Excel.Files import Files
import openpyxl
from openpyxl import Workbook
import csv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.support.relative_locator import locate_with
from selenium.webdriver.support.select import Select
@task
def notary():
    """FILL NOTARY ADVOCATE Details"""
    url = "https://botsdna.com/notaries/"
    driver = webdriver.Chrome()
    driver.get(url)
    # browser.goto("https://botsdna.com/notaries/")
    # page = browser.page()
    options = ["VIZIANAGARAM ","VISAKHAPATNAM ","EAST GODAVARI ","WEST GODAVARI ","KRISHNA ","GUNTUR ","PRAKASAM ","NELLORE ","CHITTOOR ","ANANTHAPURAM ","KADAPA ","KURNOOL ","DONE BY GOVT"]
    count=0
    for i in range(5,578):
        data = get_data(i)
        if data[0]==None:
            count+=1
        else:
            driver = webdriver.Chrome()
            driver.get(url)
            notary_form = driver.find_element(By.ID, "notary")
            notary_form.send_keys(data[0])
            area_form = driver.find_element(By.ID, "area")
            area_form.send_keys(data[1])
            select_field = driver.find_element(By.ID, "DIST")
            select = Select(select_field)
            select.select_by_index(count)
            submit_button = driver.find_element(By.XPATH,"/html/body/center/table[2]/tbody/tr[4]/td[2]/input")
            submit_button.click()
            time.sleep(2)
            tansaction_number = driver.find_element(By.ID,"TransNo").text
            workbook = openpyxl.load_workbook('AP-ADVOCATES.xlsx')
            sheet = workbook.active
            sheet.cell(row=i, column=4, value = tansaction_number)
            workbook.save('AP-ADVOCATES.xlsx')
            workbook.close()
            driver.quit()

def get_data(row_value):
    """Get the data from the excel file"""
    wb = openpyxl.load_workbook("AP-ADVOCATES.xlsx")
    sheet = wb.active
    advocate_name = sheet.cell(row=row_value, column=2).value
    area = sheet.cell(row=row_value, column=3).value
    values = [advocate_name,area]
    wb.close()
    return  values

    
    
