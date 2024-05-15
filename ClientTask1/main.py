import openpyxl
from openpyxl.styles import Font, NamedStyle, Border, Side
import csv
import pandas as pd
import re
import datetime

class Order:
    def __init__(self):
        pass
    
    def num_to_char(self, n):
        return chr(64 + n)
    
    def create_output_file(self, filename="Output.xlsx"):
        """Create a new output file with a default name 'Output.xlsx'."""
        try:
            # Load the template workbook
            source_workbook = openpyxl.load_workbook('output\Template.xlsx')

            # Get the active sheet
            active_sheet = source_workbook.active

            # Create a new workbook
            new_workbook = openpyxl.Workbook()

            # Add a new sheet to the new workbook with the same name as the active sheet
            new_sheet = new_workbook.active
            new_sheet.title = active_sheet.title

            # Copy the cells from the active sheet to the new sheet
            for row in active_sheet.iter_rows():
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                new_sheet.append(new_row)

            # Save the new workbook to a file
            new_workbook.save(f'output\{filename}')
        except Exception as e:
            print(f"Error creating output file: {e}")

    # def convert_to_xlsx_file(self, filename="Consolidated Order Statement 10022024.xls"):
    #     """Convert a .xls file to .xlsx file for easier formatting."""
    #     try:
    #         x2x = XLS2XLSX(f"input/{filename}")
    #         x2x.to_xlsx(f"Order_Statement.xlsx")
    #     except Exception as e:
    #         print(f"Error converting to xlsx file: {e}")


    def csv_to_xlsx_file(self, filename = "input\input.csv"):
        """Convert a .csv file to .xlsx file for easier formatting."""
        try:
            df = pd.read_csv('input\input.csv')

            # Write the DataFrame to an XLSX file
            df.to_excel('input\Input.xlsx', index=False)
            return "Task completed"
        except Exception as e:
            return f"Error converting CSV to XLSX: {e}"
            

    def extract_integer(self, string):
        """Extract integer from a string."""
        try:
            match = re.search(r'(\d+)', string)
            if match:
                return int(match.group(1))
            else:
                return None
        except Exception as e:
            print(f"Error extracting integer: {e}")

    def get_data(self):
        """Retrieve data from the input file."""
        try:
            vegetable_names = []
            unit_quant = []
            veg_quant = {}

            # Create the output file
            self.create_output_file()
            csv_xlsx = self.csv_to_xlsx_file()
            
            if csv_xlsx == "Task completed":
                wb = openpyxl.load_workbook(f'input\Input.xlsx')
                ws = wb.active
            else:
                try:
                    wb = openpyxl.load_workbook(f'input\Input.xlsx')
                    ws = wb.active
                except Exception as e:
                    raise FileNotFoundError("Unable to locate the Input file.")            

            # Store data in lists
            col_a = [cell.value for cell in ws['A'][1:] if cell.value is not None]
            col_q = [cell.value for cell in ws['Q'][1:] if cell.value is not None]
            col_r = [cell.value for cell in ws['R'][1:] if cell.value is not None]
            col_y = [cell.value for cell in ws['Y'][1:] if cell.value is not None]
            
            col_y = list(set(col_y))
            

            number_of_clients = list(set(col_a))  # remove duplicates from list

            for vegetable in col_r:
                veg_name = vegetable.split("|")[0].strip()
                veg_quant[veg_name] = vegetable.split("-")[-1].strip()
                unit_quant.append(vegetable.split("-")[-1].strip())
                vegetable_names.append(veg_name)

            # Get unique vegetables
            vegetables = list(set(vegetable_names))
            sorted_vegetables = sorted(vegetables)

            # Create template
            self.create_template(sorted_vegetables, veg_quant, number_of_clients)

            # Add data
            self.add_data(number_of_clients)

        except Exception as e:
            print(f"Error retrieving data: {e}")

    def create_template(self, vegetables, uom, clients):
        """Create a template excel file with headers."""
        try:
            unit = ""
            wb = openpyxl.load_workbook(f'output/Output.xlsx')
            ws = wb.active
            current_date = datetime.date.today()
            # Format the current date in the DD/MM/YYYY format
            formatted_date = current_date.strftime("%d/%m/%Y")
            ws.cell(row=3, column=1, value=f"Date : {formatted_date}")

            

            named_style = NamedStyle(name="bordered")
            named_style.border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))

            ws["A1"].font = Font(size=18, bold=True)
            ws["A2"].font = Font(size=18, bold=True)
            ws["A3"].font = Font(size=18, bold=True)

            ws.cell(row=5, column=1, value="Sr. No.").style = named_style
            ws.cell(row=5, column=2, value="Product").style = named_style
            ws.cell(row=5, column=3, value="UOM").style = named_style
            ws.cell(row=5, column=4, value="Qty.").style = named_style

            ws["A5"].font = Font(size=14, bold=True)
            ws["B5"].font = Font(size=14, bold=True)
            ws["C5"].font = Font(size=14, bold=True)
            ws["D5"].font = Font(size=14, bold=True)

            for i in range(len(clients)):
                ws.cell(row=5, column=5+i, value=clients[i]).style = named_style
                ws.cell(row=5, column=5+i).font = Font(size = 14, bold=True)
            # cell_value = len(clients)-1+5
            # print(self.num_to_char(cell_value))
            # ws.merge_cells(f'A1:{self.num_to_char(26)}1')
            
            

            for i in range(len(vegetables)):
                if "Gram" in uom[vegetables[i]] or "gram" in uom[vegetables[i]] or "Kg" in uom[vegetables[i]] or "kg" in uom[vegetables[i]]:
                    unit = "Kg"
                elif "Pc" in uom[vegetables[i]] or "pc" in uom[vegetables[i]]:
                    unit = "Pc"
                ws.append([i+1, vegetables[i], unit])

                ws.cell(row = i+6, column = 1).style = named_style
                ws.cell(row = i+6, column = 2).style = named_style
                ws.cell(row = i+6, column = 3).style = named_style
                ws.cell(row = i+6, column = 1).font = Font(size = 14)
                ws.cell(row = i+6, column = 2).font = Font(size = 14)
                ws.cell(row = i+6, column = 3).font = Font(size = 14)
                

            ctr = 1 
            max_col = ws.max_column
            max_rows = ws.max_row
            for j in range(1,max_rows+1):
                if j<=5:
                    ctr += 1
                elif j>5:
                    for i in range(5,max_col+1):
                        ws.cell(row = ctr, column = i, value = None).style = named_style
                    ctr+=1

            wb.save(f'output/Output.xlsx')
        except Exception as e:
            print(f"Error creating template: {e}")

    def add_data(self, clients):
        """Add data to existing work book."""
        try:
            col_counter = 1
            row_counter = 1

            named_style = NamedStyle(name="new_borders")
            named_style.border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))
            
            wb1 = openpyxl.load_workbook('input\Input.xlsx')
            ws1 = wb1.active

            wb2 = openpyxl.load_workbook(f'output/Output.xlsx')
            ws2 = wb2.active

            for row in ws1.iter_rows(values_only=True):
                if row[0] in clients:
                    for col_idx, cell_value in enumerate(ws2[5], start=1):  # Iterate over cells in the first row
                        if cell_value.value == row[0]:  # Check if the cell value matches the client identifier
                            if row[24]!=None:
                                substrings = row[24].split()
                                # Extract the first letter of each substring
                                initials = [substring[0] for substring in substrings]
                                # Join the initials into a single string
                                initials_string = ''.join(initials)
                                ws2.cell(row = 4, column = col_idx).style = named_style
                                ws2.cell(row = 4, column = col_idx, value = initials_string.upper()).font = Font(size = 14, bold=True)

                            col_counter = col_idx  # Update the column counter with the column index
                            break
                    
                    for row_value in ws2.iter_rows(values_only=True):
                        if row_counter<=5:
                            row_counter += 1
                        elif any(cell == row[17].split("|")[0].strip() for cell in row_value):
                            row_of_veg = row_counter
                            break
                        else:
                            row_counter += 1
                    qty = self.extract_integer(row[17].split("-")[-1].strip())
                    if qty is not None:
                        if "Gram" in row[17].split("-")[-1].strip() or "gram" in row[17].split("-")[-1].strip():
                            qty = float(qty/1000)
                            qty *= row[16]
                        else:
                            qty *= row[16]
                        ws2.cell(row=row_of_veg, column=col_counter, value=qty).font = Font(size = 14)

                    wb2.save(f'output/Output.xlsx')
                row_counter = 1
            row_counter = 1
            total = 0

            for row in ws2.iter_rows(values_only=True):
                if row_counter<=5:
                    row_counter+=1
                    pass
                else:
                    for i in range(4,len(row)):
                        if row[i]  != None:
                            total += float(row[i])
                    ws2.cell(row=row_counter, column=4).style = named_style
                    ws2.cell(row=row_counter, column=4, value=total).font = Font(size =14)
                    row_counter+=1
                    total = 0
            wb2.save(f'output/Output.xlsx')

        except Exception as e:
            print(f"Error adding data: {e}")

    #Only for testing purposes does not add any meaning to the application.
    def testing(self):
        try:
            wb1 = openpyxl.load_workbook('Order_Statement.xlsx')
            ws1 = wb1.get_sheet_by_name("Input File")

            for row in ws1.iter_rows(values_only=True):
                print(row)
                break
        except Exception as e:
            print(f"Error testing: {e}")

if __name__ == "__main__":
    order = Order()
    order.get_data()
