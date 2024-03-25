from robocorp.tasks import task
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from io import BytesIO
import datetime
import os
from PIL import Image
import io
import requests


@task
def minimal_task():
    driver = webdriver.Chrome()
    driver.get("https://www.customfit.ai/blog/uncovering-indias-top-100-d2c-brands-what-you-need-to-know")
    
    # wb = Workbook()
    # ws = wb.active
    
    # # Add column headers to the first row of the sheet
    # ws["A1"] = "Brand Name"
    # ws["B1"] = "Website"

    # wb.save("D2C.xlsx")  # Save the workbook with the specified filename
    # wb.close()

    all_companies = driver.find_elements(By.XPATH,"//div[2]/div[3]/div/h3")
    wb = openpyxl.load_workbook("D2C.xlsx")
    ws = wb.get_sheet_by_name("Sheet")
    for company in all_companies: 
        last_row = ws.max_row + 1   
        link = company.find_element(By.TAG_NAME,"a")
        link = link.get_attribute("href")
        name = company.find_element(By.TAG_NAME,"a").text

        ws.cell(row=last_row, column=1, value=name)
        ws.cell(row=last_row, column=2, value=link)
        wb.save("D2C.xlsx")
    
